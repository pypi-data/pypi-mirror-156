# -*- coding: utf-8 -*-
########################################################################
# <LUXPY: a Python package for lighting and color science.>
# Copyright (C) <2017>  <Kevin A.G. Smet> (ksmet1977 at gmail.com)
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
#########################################################################
"""
Module with utility functions and parameters
============================================

 :_PKG_PATH: absolute path to luxpy package
 
 :_SEP: operating system operator
 
 :_EPS: = 7./3 - 4./3 -1 (machine epsilon)
 
 :np2d(): Make a tuple, list or array at least a 2D numpy array.

 :np2dT(): Make a tuple, list or array at least a 2D numpy array and tranpose.

 :np3d(): Make a tuple, list or array at least a 3D numpy array.

 :np3dT(): Make a tuple, list or array at least a 3D numpy array 
           and tranpose (swap) first two axes.

 :put_args_in_db():  | Takes the args with not-None input values of a function 
                       and overwrites the values of the corresponding keys 
                       in dict db.
                     | See put_args_in_db? for more info.

 :vec_to_dict(): Convert dict to vec and vice versa.

 :getdata(): Get data from csv-file or convert between pandas dataframe
             and numpy 2d-array.

 :dictkv(): Easy input of of keys and values into dict 
            (both should be iterable lists).

 :OD(): Provides a nice way to create OrderedDict "literals".

 :meshblock(): | Create a meshed block.
               | (Similar to meshgrid, but axis = 0 is retained) 
               | To enable fast blockwise calculation.

 :aplit(): Split ndarray data on (default = last) axis.

 :ajoin(): Join tuple of ndarray data on (default = last) axis.

 :broadcast_shape(): | Broadcasts shapes of data to a target_shape. 
                     | Useful for block/vector calculations when numpy fails 
                       to broadcast correctly.

 :todim(): Expand x to dimensions that are broadcast-compatable 
           with shape of another array.
           
 :write_to_excel(): Write a DataFrame to existing an Excel file into specific Sheet.
 
 :show_luxpy_tree(): Show luxpy folder structure
 
 :is_importable(): Check if a module is importable / loaded and if it doesn't exist installing it using subprocess
 
 :get_function_kwargs(): Get dictionary of a function's keyword arguments and their default values. 

 :profile_fcn(): Profile or time a function fcn.

 :unique(): Get unique elements from array.
 
 :save_pkl(): save object in pickle file
 
 :load_pkl(): load object in pickle file
 
===============================================================================
"""
#------------------------------------------------------------------------------
# Package imports:
# Core:
import os
import warnings
import subprocess
import importlib
import time
import cProfile
import pstats
import io
import pickle
from collections import OrderedDict as odict
from mpl_toolkits.mplot3d import Axes3D
__all__ = ['odict','Axes3D']

# other:
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import scipy as sp

__all__ += ['np','pd','plt','sp']


#------------------------------------------------------------------------------
# os related utility parameters:
_PKG_PATH = os.path.dirname(__file__);""" Absolute path to package """ 
_PKG_PATH = _PKG_PATH[:_PKG_PATH.find("utils")-1]
_SEP = os.sep; """ Operating system separator """
__all__ += ['_PKG_PATH','_SEP']


#------------------------------------------------------------------------------
# set some general utility parameters:
_EPS = 7./3 - 4./3 -1; """ Machine epsilon """
__all__+=['_EPS']


#------------------------------------------------------------------------------
from .folder_tree import tree
__all__ += ['np2d','np3d','np2dT','np3dT',
           'put_args_in_db','vec_to_dict',
           'getdata','dictkv','OD','meshblock','asplit','ajoin',
           'broadcast_shape','todim','write_to_excel','show_luxpy_tree',
           'is_importable','get_function_kwargs','profile_fcn','unique',
           'save_pkl', 'load_pkl']

##############################################################################
# Start function definitions
##############################################################################

#------------------------------------------------------------------------------
def np2d(data):
    """
    Make a tuple, list or numpy array at least a 2D numpy array.
    
    Args:
        :data: 
            | tuple, list, ndarray
        
    Returns:
        :returns:
            | ndarray with .ndim >= 2
    """
    if isinstance(data, np.ndarray):# assume already atleast_2d when nd.array (user has to ensure input is an array)
        if (len(data.shape)>=2):
            return data   
        else:
            return np.atleast_2d(data)
    else:
        return np.atleast_2d(np.array(data))


def np2dT(data):
    """
    Make a tuple, list or numpy array at least a 2D numpy array and transpose.
    
    Args:
        :data: 
            | tuple, list, ndarray
        
    Returns:
        :returns: 
            | ndarray with .ndim >= 2 and with transposed axes.
    """
    if isinstance(data, np.ndarray):# assume already atleast_2d when nd.array (user has to ensure input is an array)
        if (len(data.shape)>=2):
            return data.T   
        else:
            return np.atleast_2d(data).T
    else:
        return np.atleast_2d(np.array(data)).T

def np3d(data):
    """
    Make a tuple, list or numpy array at least a 3d numpy array.
    
    Args:
        :data: 
            | tuple, list, ndarray
        
    Returns:
        :returns: 
            | ndarray with .ndim >= 3
    """
    if isinstance(data, np.ndarray):# assume already atleast_3d when nd.array (user has to ensure input is an array)
        if (len(data.shape)>=3):
            return data   
        else:
            return np.expand_dims(np.atleast_2d(data),axis=0)
    else:
        return np.expand_dims(np.atleast_2d(np.array(data)),axis=0)
    
def np3dT(data): # keep last axis the same
    """
    Make a tuple, list or numpy array at least a 3d numpy array and transposed first 2 axes.
    
    Args:
        :data: 
            | tuple, list, ndarray
        
    Returns:
        :returns: 
            | ndarray with .ndim >= 3 and with first two axes 
            | transposed (axis=3 is kept the same).
    """
    if isinstance(data,np.ndarray):# assume already atleast_3d when nd.array (user has to ensure input is an array)
        if (len(data.shape)>=3):
            return data.transpose((1,0,2))
        else:
            return np.expand_dims(np.atleast_2d(data),axis=0).transpose((1,0,2))
    else:
        return np.expand_dims(np.atleast_2d(np.aray(data)),axis=0).transpose((1,0,2))



#------------------------------------------------------------------------------
def put_args_in_db(db, args):
    """
    Takes the args with not-None input values of a function and overwrites 
    the values of the corresponding keys in dict db.
    | (args are collected with the built-in function locals(), 
    | See example usage below)
    
    Args:
        :db: 
            | dict
    
    Returns:
        :returns: 
            | dict with the values of specific keys overwritten by the 
            |      not-None values of corresponding args of a function fcn.
    
    Example usage:
        | _db = {'c' : 'c1', 'd' : 10, 'e' : {'e1':'hello', 'e2':1000}}
        |
        | def test_put_args_in_db(a, b, db = None, c = None,d = None,e = None):
        | 
        |    args = locals().copy()  # get dict with keyword input arguments to 
        |                             # function 'test_put_args_in_db'
        |     
        |     db = put_args_in_db(db,args) # overwrite non-None args in db copy.
        |     
        |    if db is not None: # unpack db for further use
        |        c,d,e = [db[x] for x in sorted(db.keys())]
        |     
        |     print(' a : {}'.format(a))
        |     print(' b : {}'.format(b))
        |     print(' db: {}'.format(db))
        |     print(' c : {}'.format(c))
        |     print(' d : {}'.format(d))
        |     print(' e : {}'.format(e))
        |     print('_db: {}'.format(_db))
 
    """
    # overwrite not-'None' input arguments in db:
    if db is not None:
        dbc = db.copy()
        dbc_keys = dbc.keys()
        for argkey in args.keys():
            if args[argkey] is not None:
                if argkey in dbc_keys:
                    dbc[argkey] =  args[argkey]
        return dbc
    else:
        return db

#------------------------------------------------------------------------------   
def vec_to_dict(vec= None, dic = {}, vsize = None, keys = None):
    """
    Convert dict to vec and vice versa.
    
    Args:
        :vec: 
            | list or vector array, optional
        :dic: 
            | dict, optional
        :vsize:
            | list or vector array with size of values of dict, optional
        :keys:
            | list or vector array with keys in dict (must be provided).
        
    Returns:
        :returns:
            | x, vsize
            |   x is an array, if vec is None
            |   x is a dict, if vec is not None
    """
    if vec is not None:
        # Put values in vec in dic:
        n = 0 # keeps track of length already read from x
        for i,v in enumerate(keys):
            dic[v] = vec[n + np.arange(vsize[i])]
            n += dic[v].shape[0] 
        return dic, vsize
    else:
        # Put values of keys in dic in vec:
        vec = []
        vsize = []
        for i,v in enumerate(keys):
            vec = np.hstack((vec, dic[v]))
            vsize.append(dic[v].shape[0])
        return vec, vsize

#--------------------------------------------------------------------------------------------------
def getdata(data, kind = 'np', columns = None, header = None, sep = ',', datatype = 'S', copy = True, verbosity = True):
    """
    Get data from csv-file 
    or convert between pandas dataframe and numpy 2d-array.
    
    Args:
        :data: 
            | - str with path to file containing data
            | - ndarray with data
            | - pandas.dataframe with data
        :kind: 
            | str ['np','df'], optional 
            | Determines type(:returns:), np: ndarray, df: pandas.dataframe
        :columns:
            | None or list[str] of column names for dataframe, optional
        :header:
            | None, optional
            |   - None: no header in file
            |   - 'infer': infer headers from file
        :sep:
            | ',' or '\t' or other char, optional
            | Column separator in data file
        :datatype':
            | 'S',optional 
            | Specifies a type of data. 
            | Is used when creating column headers (:column: is None).
            |   -'S': light source spectrum
            |   -'R': reflectance spectrum
            |   or other.   
        :copy:
            | True, optional
            | Return a copy of ndarray if kind == 'np', or copy of pd.DataFrame if kind == 'df'
        :verbosity:
            | True, False, optional
            | Print warning when inferring headers from file.
    
    Returns:
        :returns:
            | data as ndarray or pandas.dataframe
      
    """
    if isinstance(data,str):
        datafile = data
        data = pd.read_csv(data,names=None,index_col = None,header = header,sep = sep)

        # Set column headers:
        if header == 'infer':
            if verbosity == True:
                warnings.warn('getdata(): Infering HEADERS from data file: {}!'.format(datafile))
                columns = data.columns
        elif (columns is None):
            data.columns = ['{}{}'.format(datatype,x) for x in range(len(data.columns))] 
        if columns is not None:
            data.columns = columns

    if isinstance(data,np.ndarray) & (kind == 'df'):
        if columns is None:
            columns = ['{}{}'.format(datatype,x)  for x in range(data.shape[1])] 
        data = pd.DataFrame(data, columns = columns)

    elif isinstance(data,pd.DataFrame) & (kind == 'np'):
        data = data.values
    elif copy == True:
        data = data.copy()
    return data


#--------------------------------------------------------------------------------------------------
def dictkv(keys=None,values=None, ordered = True): 
    """
    Easy input of of keys and values into dict.
    
    Args:
        :keys: 
            | iterable list[str,...] of keys
        :values:
            | iterable list[...,..., ] of values
        :ordered:
            | True, False, optional
            | True: creates an ordered dict using 'collections.OrderedDict()'
    Returns:
        :returns:
            | (ordered) dict
    """
    if ordered is True:
        return odict(zip(keys,values))
    else:
        return dict(zip(keys,values))
    
class OD(object):
    """
    This class provides a nice way to create OrderedDict "literals".
    
    """
    def __getitem__(self, slices):
        if not isinstance(slices, tuple):
            slices = slices,
        return odict((slice.start, slice.stop) for slice in slices)
# Create a single instance; we don't ever need to refer to the class.
OD = OD()

#--------------------------------------------------------------------------------------------------
def meshblock(x,y):
    """
    Create a meshed block from x and y.
    
    | (Similar to meshgrid, but axis = 0 is retained).
    | To enable fast blockwise calculation.
    
    Args: 
        :x: 
            | ndarray with ndim == 2
        :y: 
            | ndarray with ndim == 2
        
    Returns:
        :X,Y: 
            | 2 ndarrays with ndim == 3 
            |   X.shape = (x.shape[0],y.shape[0],x.shape[1])
            |   Y.shape = (x.shape[0],y.shape[0],y.shape[1])
    """
    Y = np.transpose(np.repeat(y,x.shape[0],axis=0).reshape((y.shape[0],x.shape[0],y.shape[1])),axes = [1,0,2])
    X = np.repeat(x,y.shape[0],axis=0).reshape((x.shape[0],y.shape[0],x.shape[1]))
    return X,Y

#---------------------------------------------------------------------------------------------------
def asplit(data):
    """
    Split data on last axis
    
    Args:
        :data: 
            | ndarray
        
    Returns:
        :returns: 
            | ndarray, ndarray, ... 
            |   (number of returns is equal data.shape[-1])
    """
    #return np.array([data.take([x],axis = len(data.shape)-1) for x in range(data.shape[-1])])
    return [data[...,x] for x in range(data.shape[-1])]


def ajoin(data):
    """
    Join data on last axis.
    
    Args:
        :data:
            | tuple (ndarray, ndarray, ...)
        
    Returns:
        :returns:
            | ndarray (shape[-1] is equal to tuple length)
    """
    if data[0].ndim == 2: #faster implementation
        return np.transpose(np.concatenate(data,axis=0).reshape((np.hstack((len(data),data[0].shape)))),(1,2,0))
    elif data[0].ndim == 1:
        return np.concatenate(data,axis=0).reshape((np.hstack((len(data),data[0].shape)))).T
    else:
        return np.hstack(data)[0]
    

#---------------------------------------------------------------------------------------------------
def broadcast_shape(data,target_shape = None, expand_2d_to_3d = None, axis0_repeats = None, axis1_repeats = None):
    """
    Broadcasts shapes of data to a target_shape.
    
    | Useful for block/vector calc. when numpy fails to broadcast correctly.
    
    Args:
        :data: 
            | ndarray 
        :target_shape: 
            | None or tuple with requested shape, optional
            |   - None: returns unchanged :data:
        :expand_2d_to_3d:
            | None (do nothing) or ..., optional 
            | If ndim == 2, expand from 2 to 3 dimensions
        :axis0_repeats:
            | None or number of times to repeat axis=0, optional
            |   - None: keep axis=0 same size
        :axis1_repeats:
            | None or number of times to repeat axis=1, optional
            |   - None: keep axis=1 same size

    Returns:
        :returns: 
            | reshaped ndarray
    """
    data = np2d(data)
    
    # expand shape along axis (useful for some functions that allow block-calculations when the data is only 2d )
    if (expand_2d_to_3d is not None) & (len(data.shape) == 2):
        data = np.expand_dims(data, axis = expand_2d_to_3d)
     
    if target_shape is not None:
        dshape = data.shape
        if dshape != target_shape:
            axis_of_v3 = len(target_shape)-1
            if (dshape[0] != target_shape[0]): # repeat along axis 0
                if axis0_repeats is None:
                    axis0_repeats = (target_shape[0]-dshape[0] + 1)
                data = np.repeat(data,axis0_repeats,axis = 0)
            if (len(target_shape)>2) & (len(data.shape)==2): # repeat along axis 1, create axis if necessary
                data = np.expand_dims(data,axis = axis_of_v3-1) # axis creation
                dshape = data.shape
                if (dshape[1] != target_shape[1]):
                    if axis1_repeats is None:
                        axis1_repeats = (target_shape[1]-dshape[1] + 1) # repititon
                    data = np.repeat(data,axis1_repeats,axis = 1)

        for i in range(2):
            if (data.shape[i] > 1) & (data.shape[i] != target_shape[i]):
                raise Exception('broadcast_shape(): Cannot match dim of data with target: data.shape[i]>1  & ...  != target.shape[i]')
    return data


def todim(x,tshape, add_axis = 1, equal_shape = False): 
    """
    Expand x to dims that are broadcast-compatable with shape of another array.
    
    Args:
        :x: 
            | ndarray
        :tshape: 
            | tuple with target shape
        :add_axis:
            | 1, optional
            | Determines where in x.shape an axis should be added
        :equal_shape:
            | False or True, optional
            | True: expand :x: to identical dimensions (speficied by :tshape:)
            
    Returns:
        :returns:
            | ndarray broadcast-compatable with tshape.
    """
    if x is None:
        return np.broadcast_arrays(x,np.ones(tshape))[0]
    else:
        x = np2d(x)
        sx = x.shape
        lsx = len(sx)
        ltshape = len(tshape)
        if (sx == tshape):
            pass
        else:
            
            if ((lsx == 1) | (sx == (1,tshape[-1])) | (sx == (tshape[-1],1))): 
                if (sx == (tshape[-1],1)):
                    x = x.T
                if lsx != ltshape:
                    x = np.expand_dims(x,0)
            elif (lsx == 2):
                if (ltshape == 3):
                    sd = np.setdiff1d(tshape, sx,assume_unique=True)
                    if len(sd) == 0:
                        ax = add_axis
                    else:
                        ax = np.where(tshape==sd)[0][0]
                    x = np.expand_dims(x,ax)
                else:
                    raise Exception("todim(x,tshape): dimensions do not match for 2d arrays.")  
            else:
                raise Exception("todim(x,tshape): no matching dimensions between 3d x and tshape.")
        if equal_shape == False:
            return x
        else:
            return np.ones(tshape)*x #make dims of x equal to those of a (tshape)
        
#------------------------------------------------------------------------------
def write_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Writes a DataFrame to an existing Excel file into a specified sheet.
    | If [filename] doesn't exist, then this function will create it.

    Args:
      :filename: 
          | File path or existing ExcelWriter
          | (Example: '/path/to/file.xlsx')
      :df: 
          | dataframe to save to workbook
      :sheet_name: 
          | Name of sheet which will contain DataFrame.
          | (default: 'Sheet1')
      :startrow: 
          | upper left cell row to dump data frame.
          | Per default (startrow=None) calculate the last row
          | in the existing DF and write to the next row...
      :truncate_sheet: 
          | truncate (remove and recreate) [sheet_name]
          | before writing DataFrame to Excel file
      :to_excel_kwargs: 
          | arguments which will be passed to `DataFrame.to_excel()`
          | [can be dictionary]

    Returns: None
    
    Notes:
        Copied from https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
def show_luxpy_tree(omit = ['.pyc','__pycache__',
                            '.txt','.dat','.csv','.npz',
                            '.png','.jpg','.md','.pdf','.ini','.log', '.rar',
                            'drivers','SDK_','dll','bak']):
    """
    Show luxpy foler tree.
    
    Args:
        :omit:
            | List of folders and file-extensions to omit.
            
    Returns:
        None
    """
    tree(_PKG_PATH, omit = omit)
    return None


#------------------------------------------------------------------------------
def is_importable(string, try_pip_install = False):
    """
    Check if string is importable/loadable. If it doesn't then try to 'pip install' it using subprocess.
    Returns None if succesful, otherwise throws and error or outputs False.
    
    Args:
        :string:
            | string with package or module name
        :try_pip_install:
            | False, optional
            | True: try pip installing it using subprocess
    
    Returns:
        :success:
            | True if importable, False if not.
    """ 
    success = importlib.util.find_spec(string) is not None
    if (not success) & (try_pip_install == True):  
        try:
            print("Trying to 'pip install {:s}' using subprocess.".format(string))
            success = subprocess.call(["pip", "install", "{:s}".format(string)])
            print("subprocess output: ", success)
            if success != 0:
                raise Exception("Tried importing '{:s}', then tried pip installing it. Please install it manually: pip install {:s}".format(string,string))  
            else:
                print("'pip install {:s}' succesful".format(string))
            success = importlib.util.find_spec(string) is not None
        except:
            success = False
            raise Exception("Tried importing '{:s}', then tried pip installing it. Please install it manually: pip install {:s}".format(string,string))   
    return success
    
#------------------------------------------------------------------------------
def get_function_kwargs(f):
    """
    Get dictionary of a function's keyword arguments and their default values. 
    
    Args:
        :f:
            | function name
    
    Returns:
        :dict:
            | Dict with the function's keyword arguments and their default values
            | Is empty if there are no defaults (i.e. f.__defaults__ or f.__kwdefaults__ are None).
    """
    kwdefs = f.__kwdefaults__
    if kwdefs is None: kwdefs = {}
    names = f.__code__.co_varnames[:f.__code__.co_argcount]
    if f.__defaults__ is not None:
        d = dict(zip(names[::-1][:len(f.__defaults__)][::-1],f.__defaults__))
    else:
        d = {}
    d.update(kwdefs)
    return d

#------------------------------------------------------------------------------
def profile_fcn(fcn, profile=True, sort_stats = 'tottime', output_file=None):
    """
    Profile or time a function fcn.
    
    Args:
        :fcn:
            | function to be profiled or timed (using time.time() difference)
        :profile:
            | True, optional
            | Profile the function, otherwise only time it.
        :sort_stats:
            | 'tottime', optional
            | Sort profile results according to sort_stats ('tottime', 'cumtime',...)
        :output_file:
            | None, optional
            | If not None: output result to output_file.
            
    Return:
        :ps:
            | Profiler output
            
    """
   
    if profile == False:
        start = time.time()
        fcn()
        dt = time.time()-start
        print('%s %f' % ('Time elapsed: ', dt))
        return dt
    else:
        pr = cProfile.Profile()
        pr.enable()
        fcn()
        pr.disable()
        s = io.StringIO()
        ps = pstats.Stats(pr, stream=s).sort_stats(sort_stats)
        ps.print_stats()
        if output_file is not None:
            with open(output_file, 'w+') as f:
                f.write(s.getvalue())
        return ps
    
#------------------------------------------------------------------------------
def unique(array, sort = True):
    """ 
    Get unique elements from array.
    
    Args:
        :array:
            | array to get unique elements from.
        :sort:
            | True, optional
            | If True: get sorted unique elements.
            
    Returns:
        :unique_array:
            | ndarray with (sorted) unique elements.
    """
    if sort:
        return np.unique(array) 
    else:
        uniq, index = np.unique(array, return_index=True)
        return uniq[index.argsort()]

#------------------------------------------------------------------------------
def save_pkl(filename, obj): 
    """ 
    Save an object in a pickle file.
    
    Args:
        :filename:
            | str with filename of pickle file.
        :obj:
            | python object to save
    
    Returns:
        :None:
    """
    with open(filename, 'wb') as handle:
        pickle.dump(obj, handle, protocol=pickle.HIGHEST_PROTOCOL)
        
def load_pkl(filename):
    """ 
    Load the object in a pickle file.
    
    Args:
        :filename:
            | str with filename of pickle file.
        
    Returns:
        :obj:
            | loaded python object
    """
    obj = None
    with open(filename, 'rb') as handle:
        obj = pickle.load(handle)
    return obj