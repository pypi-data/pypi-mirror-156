# coding=utf-8
import os
import logging
import hmac
import base64
import random
import time
from pathlib import Path
from datetime import timedelta, datetime, date
from string import ascii_letters
import zlib
import hashlib, string
import json
from bson.objectid import ObjectId
import operator
import re
import traceback

_chars = string.printable[:87] + '_' + string.printable[90:95]
to_ObjectId = lambda a: ObjectId(a) if type(a) != ObjectId else a
to_python = lambda s: json.loads(s)
to_json = lambda obj: json.dumps(obj, ensure_ascii=False, sort_keys=True)
fix_path = lambda path: Path(path).as_posix()
traceback = traceback

num_or_alpha = re.compile("^(?!\d+$)[\da-zA-Z_]{5,10}$")                            # 仅数字和字母组合，不允许纯数字,长度5~10
startwith_alpha = re.compile("^[a-zA-Z]{5,10}")                                     # 仅允许以字母开头,长度5~10
lllegal_char = re.compile('^[_a-zA-Z0-9\u4e00-\u9fa5]+$')                           # 仅允许中文、英文、数字,_下划线。但不允许非法字符
email_re = re.compile("^.+\\@(\\[?)[a-zA-Z0-9\\-\\.]+\\.([a-zA-Z]{2,3}|[0-9]{1,3})(\\]?)$")    # 是否是邮箱
phone_re = re.compile("^(0|86|17951)?(13|14|15|16|17|18)[0-9]{9}$")                 # 11位手机号
password_normal = re.compile("^(?:(?=.*)(?=.*[a-z])(?=.*[0-9])).{6,12}$")      # 一般密码 密码必须包含字母，数字,任意字符，长度6~12
password_strong = re.compile("^(?:(?=.*[A-Z])(?=.*[a-z])(?=.*[0-9])).{6,12}$")      # 强密码 密码必须包含大小写，数字,任意字符，长度6~12

def random_int(length=6):
    """生成随机的int 长度为length"""
    return ("%0" + str(length) + "d") % random.randint(int("1" + "0" * (length - 1)), int("9" * length))

def gen_random_str(length=6, chars=_chars):
    return ''.join(random.choice(chars) for i in range(length))

def gen_random_sint(length=6):
    """
    获取字符串加数字的随机值
    :param length:
    :return:
    """
    return "".join(random.choice(string.hexdigits) for i in range(length))

def random_string(length=6):
    """生成随机字符串"""
    return ''.join(random.choice(ascii_letters) for i in range(length))

def gen_hmac_key():
    """随机生成长度32位密文"""
    s = str(ObjectId())
    k = gen_random_str()
    key = hmac.HMAC(k, s).hexdigest()
    return key

def enbase64(s):
    """
    编码
    :param s:
    :return:
    """
    if type(s) == bytes:
        return base64.b64encode(s)
    else:
        s = s.encode('utf-8')
        return base64.b64encode(s).decode('utf-8')

def debase64(s):
    """
    解码
    :param s:
    :return:
    """
    bytes_types = (bytes, bytearray)
    return base64.b64decode(s) if isinstance(s, bytes_types) else base64.b64decode(s).decode()


class TypeConvert(object):
    """类型转换类, 处理参数"""
    MAP = {int: int,
           float: float,
           bool: bool,
           str: str}

    STR2TYPE = {"int": int,
                "integer": int,
                "string": str,
                "str": str,
                "bool": bool,
                "float": float,
                "list": list,
                "dict": dict,
                "json": json.loads}

    @classmethod
    def apply(cls, obj, raw):
        try:
            tp = type(obj)
            if tp in TypeConvert.MAP:
                return TypeConvert.MAP[tp](raw)
            return obj(raw)
        except Exception as e:
            logging.error("in TypeConvert.apply %s, obj: %s, raw: %s" % (e, obj, raw))
            return None

    @classmethod
    def convert_params(cls, _type, value):
        if _type in ["int", "integer"] and not value:
            value = 0
        try:
            tp = TypeConvert.STR2TYPE[_type]
            return tp(value)
        except Exception as e:
            raise e

def calculate_age(ts):
    """计算年龄"""
    if ts == -1:
        return -1
    born = datetime.fromtimestamp(ts)
    today = date.today()
    return today.year - born.year - ((today.month, today.day) < (born.month, born.day))


def now():
    return int(time.time())

def oid_to_date(oid):
    return int_to_date_string(int(str(oid)[:8], 16))


def int_to_date_string(ts, fm=False):
    fm = fm if fm else "%Y-%m-%d %H:%M"
    try:
        if not ts:
            ts = 0
        return datetime.fromtimestamp(ts).strftime(fm)
    except:
        return datetime.fromtimestamp(time.time()).strftime(fm)

def str2timestamp(date_string, fm=False):
    fm = fm if fm else "%Y-%m-%d"
    return int(time.mktime(time.strptime(date_string, fm)))
    # return int(time.mktime(datetime.strptime(date_string, format).timetuple()))

def timestamp2str(ts, fm=False):
    return int_to_date_string(ts, fm)

def get_yesterday_midnight():
    # 获取昨天的午夜时间戳
    return get_today_midnight() - 86400

def get_today_midnight():
    # 获取今天的0时时间戳
    now = int(time.time())
    return now - now % 86400 - 3600 * 8

def get_today_lasttime():
    # 获取今天23：59：59时间戳
    now = int(time.time())
    return now - now % 86400 - 3600 * 8 + 24 * 3600 - 1


def get_delta_day(day=1):
    """
    获取n天后的时间
    :param day:
    :return:
    """

    now = datetime.now()

    # 当前日期
    now_day = now.strftime('%Y-%m-%d %H:%M:%S')

    # n天后
    delta_day = (now + timedelta(days=int(day))).strftime("%Y-%m-%d %H:%M:%S")
    return delta_day

def get_ts_from_object(s):
    if len(s) == 24:
        return int(s[:8], 16)
    return 0

def compress_obj(dict_obj, compress=True):
    """反序列化dict对象"""
    dict_obj = {"$_key_$": dict_obj} if not isinstance(dict_obj, dict) else dict_obj
    if compress:
        return zlib.compress(to_json(dict_obj))
    return to_json(dict_obj)

def uncompress_obj(binary_string, compress=True):
    """反序列化dict对象"""
    if compress:
         dict_obj = to_python(zlib.decompress(binary_string))
    else:
        dict_obj = to_python(binary_string)

    if "$_key_$" in dict_obj:
        return dict_obj["$_key_$"]
    else:
        return dict_obj

def get_mod(uid, mod=10):
    return int(uid) % mod

def gen_salt(len=6):
    return ''.join(random.sample(string.ascii_letters + string.digits, len))

def gen_salt_pwd(salt, pwd):
    """
    MD5带密码加密，返回十六进制字符串
    """
    return hashlib.md5((str(salt) + str(pwd)).encode("utf-8")).hexdigest()

def md5(s):
    """
    简单md5，返回十六进制字符串
    """
    s = s.encode("utf-8") if type(s) != bytes else s
    return hashlib.md5(s).hexdigest()

def pparams(request, params):
    print_params = {}
    for k, v in params.items():
        if k not in request.files:
            print_params[k] = v
        else:
            try:
                if "body" in v:
                    tmp = {}
                    tmp.update(v)
                    tmp["body"] = "%d bytes" % len(v["body"])
                    print_params[k] = tmp
                else:
                    print_params[k] = v
            except Exception as e:
                print_params[k] = {}

    return print_params

def version_cmp(version1, version2):
    """比较系统版本号
    v1 > v2 1
    v1 = v2 0
    v1 < v2 -1
    v1: 用户使用的版本
    v2：最新上线的版本
    """

    def normalize(v):
        return [int(x) for x in re.sub(r'(\.0+)*$', '', v).split(".")]

    return operator.gt(normalize(version2), normalize(version1))

def _find_option_with_arg(argv, short_opts=None, long_opts=None):
    """Search argv for options specifying short and longopt alternatives.

    Returns:
        str: value for option found
    Raises:
        KeyError: if option not found.

    Example：
        config_name = _find_option_with_arg(short_opts="-F", long_opts="--config")
    """
    for i, arg in enumerate(argv):
        if arg.startswith('-'):
            if long_opts and arg.startswith('--'):
                name, sep, val = arg.partition('=')
                if name in long_opts:
                    return val if sep else argv[i + 1]
            if short_opts and arg in short_opts:
                return argv[i + 1]
    raise KeyError('|'.join(short_opts or [] + long_opts or []))

def check2json(data):
    if isinstance(data, (list, tuple)):
        for index, item in enumerate(data):
            data[index] = check2json(item)
        return data
    elif isinstance(data, dict):
        for key, value in data.items():
            data[key] = check2json(value)
        return data
    elif isinstance(data, ObjectId):
        return str(data)
    else:
        return data


def verify_username(username, min: int =4, max: int =10):
    """
    校验用户名
    :param username: 待验证用户名
    :param min: 最小长度
    :param max: 最大长度
    :return:
    """
    re_str1 = re.compile("^[a-zA-Z](?!\d+$)[\da-zA-Z_]")      # 非数字开头
    re_str2 = re.compile("^(?!\d+$)[\da-zA-Z_]{%d,%d}$" % (min, max))      # 非数字开头，字母与数字组合，长度4~10

    return re.search(re_str1, username) and re.search(re_str2, username)

def verify_nickname(nickname, min: int = 4, max: int = 10):
    """
    校验昵称
    :param nickname: 待验证昵称
    :param min: 最小长度
    :param max: 最大长度
    :return:
    """
    # 允许中文、字母、数字组合，禁止非法字符,长度4~10
    lllegal_char = re.compile('^[_a-zA-Z0-9\u4e00-\u9fa5]{%d,%d}$' % (min, max))
    return re.search(lllegal_char, nickname)

def verify_password(password, min: int = 4, max: int = 10, strong_level=1):
    """
    校验密码
    :param password: 待验证密码
    :param min: 最小长度
    :param max: 最大长度
    :param strong: 密码强度等级
            以下所有规则下的密码都禁止包含空格、表情、汉字
            1、一般密码：符合长度且不能是重复或者连续的数字
            2、正常密码：必须包含字母、数字
            3、强密码：必须包含字母（含大写或小写）、数字
            4、强密码：必须包含字母、数字、特殊英文字符
            5、史上最强密码：必须包含字母（含大小写）、数字、特殊英文字符

    :return:
    """
    password_normal = re.compile("^(?:(?=.*)(?=.*[A-Za-z])(?=.*[0-9])).{%d,%d}$" % (min, max)) # 一般密码 密码必须包含字母(大写或小写)，数字，长度6~12
    password_strong = re.compile("^(?:(?=.*[A-Z])(?=.*[a-z])(?=.*[0-9])).{%d,%d}$" % (min, max))  # 强密码 密码必须含字母（含大写），数字，任意字符，长度6~12
    special_char = ["~", "!", "@", "#", "$", "%", "^", "&", "*", "(", ")", "_", "+", "=", "|", ".", ",", "?", "{", "}", "[", "]", ":"]

    if strong_level == 1:
        if len(password) < min or len(password) > max:
            result = False
        elif is_simple(password) == True:
            result = False
        else:
            result = True
    elif strong_level == 2:
        result = re.search(password_normal, password)
    elif strong_level == 3:
        result = re.search(password_strong, password)
    elif strong_level == 4:
        result = re.search(password_normal, password)
        if result:
            result = any(c in special_char for c in password)
    elif strong_level == 5:
        result = re.search(password_strong, password)
        if result:
            result = any(c in special_char for c in password)
    else:
        result = False

    if not result:
        return False

    # 禁止中文、标签
    if extract_chinese(password):
        return False

    # 禁止非法全角字符
    if extract_illegal_char(password):
        return False

    # 是否含表情
    if password != replace_emoji(password):
        return False

    return True

def verify_phone(phone):
    """
    校验手机号
    :param phone: 待验证phone
    :param min: 最小长度
    :param max: 最大长度

    :return:
    """
    phone_re = re.compile("^(0|86|17951)?(13|14|15|16|17|18)[0-9]{9}$")                 # 11位手机号
    return re.search(phone_re, phone)

def verify_email(email, min: int = 7, max: int = 50):
    """
    校验邮箱
    :param phone: 待验证邮箱
    :param min: 最小长度
    :param max: 最大长度

    :return:
    """
    # 是否是邮箱,且长度在7~50之间
    # if not re.match(r'^[0-9a-zA-Z_]{0,19}@[0-9a-zA-Z]{1,13}\.[com,cn,net]{1,3}$', email):
    #     return False
    # else:
    #     return True
    email_re = re.compile("^.+\\@(\\[?)[a-zA-Z0-9\\-\\.]+\\.([a-zA-Z]{2,3}|[0-9]{1,3})(\\]?){%d,%d}$" % (min, max))
    return re.search(email_re, email)

def verify_ip(ip, IPv6=True):
    """
    校验IP
    :param phone: 待验证ip
    :param min: 最小长度
    :param max: 最大长度

    :return:
    """
    ipv4_re = r"^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$"    # IPV4
    ipv6_re = r"^(?:[A-F0-9]{1,4}:){7}[A-F0-9]{1,4}$"    # IPV6
    if IPv6:    # 大小写不敏感
        return re.search(ipv4_re, ip, re.I) or re.search(ipv6_re, ip)
    else:
        return re.search(ipv4_re, ip)

def extract_chinese(str):
    """
    提取中文
    :param str:
    :return:
    """

    regex_str = ".*?([\u4E00-\u9FA5]+).*?"
    match_obj = re.findall(regex_str, str)
    return match_obj

def extract_illegal_char(str):
    """
    提取非法半角字符，空格
    :param str:
    :return:
    """
    regex_str = "([！￥……（）——，。、？；：【】「」《》“”‘'  ]+).*?"
    match_obj = re.findall(regex_str, str)
    return match_obj

def extract_illegal_char2(str):
    """
    性能较差
    判断是否含有特殊非法字符
    :param str:
    :return:
    """
    # 所有非法字符
    all_illegal_char = range(0x0,0x10ffff)      # 总数110万个
    for s in str:
        if s in all_illegal_char:               # 判断一个字符大约需要20ms
            return False
    return True

def change_emoji(char_str):
    """
    将表情转为字符
    :param str: 123🆚456
    :param result: 123:VS_button:456

    :param str: 表情🏷
    :param result: 表情:label:

    :return:
    """
    import emoji
    return emoji.demojize(char_str)

def replace_emoji(char_str, rep_str=''):
    """
    将字符串中的表情替换为指定的字符，默认替换为空，即过滤表情
    常用语用户名、昵称注册等禁止表情的输入
    :param char_str:
    :param rep_str:
    :return:
    """
    try:
        co = re.compile(u'[\U00010000-\U0010ffff]')
    except re.error:
        co = re.compile(u'[\uD800-\uDBFF][\uDC00-\uDFFF]')
    return co.sub(rep_str, char_str)

def is_simple(password):
    """
    判断字符串密码是否为简单重复或者纯数字或连续的数字
    若是，返回True
    若不是，返回False
    :param password:
    :return:
    """
    if password.isdigit():
        password = [int(n) for n in password]
        if len(set(password)) == 1:
            return True

        for index, a in enumerate(password[: -1]):
            b = password[index+1]
            if (b - a) != 1:
                return False
        return True
    else:
        password = [n for n in password]
        if len(set(password)) == 1:
            return True
        else:
            return False

def transfer_str(str):
    """
    mongo 正则匹配 转义
    :param str:
    :return:
    """
    new_str = ""
    special = ['/', '^', '$', '*', '+', '?', '.', '(', ')']
    for c in str:
        if c in special:
            new_str += '\\'
        new_str += c
    return new_str


def supper_format(filename, format_list):
    """
    判断文件是否在指定的格式内
    :param filename: abc.e.f.tar.gz
    :param format_list: ["rar", "tar", "tar.gz"]
    :return: True
    """
    return any(filter(lambda x: filename.lower().endswith(x.lower()), format_list))

def get_filename_format(filename, format_list):
    """
    截取文件名和格式名
    针对压缩包名称较为复杂的情况
    :param filename: abc.e.f.tar.gz
    :param format_list: ["rar", "tar", "tar.gz"]
    :return: abc.e.f
    """
    if "." not in filename:
        return filename, ""

    exists_format = False

    for i in format_list:
        if not i.startswith("."):
            suffix = "." + i
        else:
            suffix = i
        if filename.lower().endswith(i):
            name, format = filename.rsplit(suffix, 1)[0], i
            exists_format = True
            break

    if exists_format:
        return name, format
    else:
        return filename.rsplit(".", 1)

def auto_rename(name, uncompress_path, n=1):
    """
    指定一个解压位置，若有同名文件夹存在，则自动重命名
    :param name:
    :param uncompress_path:
    :param n:
    :return:
    """
    while n < 100:
        if n > 1:
            dst_path = os.path.join(uncompress_path, name + "-" + str(n))
        else:
            dst_path = os.path.join(uncompress_path, name)

        if os.path.exists(dst_path):
            n += 1
            return auto_rename(name, uncompress_path, n)
        else:
            if n > 1:
                return name + "-" + str(n), dst_path
            else:
                return name, dst_path
    raise

def undecompress(compress_path, uncompress_path):
    """
    解压包
    :param compress_path: 压缩包目前所在目录: /code/123.zip
    :param uncompress_path: 需要解压至该目录:/code/newname
    :param decompress_format: 支持的解压格式列表["zip", "rar", "7z", "tar", "tbz2", "tgz", "tar.bz2", "tar.gz", "tar.xz", "tar.Z"]
    :return:
    """
    import patoolib, shutil
    result = 1
    try:
        # 如解压目标路径里存在文件，会提示是否覆盖，一直等待输入，导致程序卡住：[Y]es, [N]o, [A]ll, n[E]ver, [R]ename, [Q]uit
        try:
            shutil.rmtree(uncompress_path)
        except:
            pass
        if not os.path.exists(uncompress_path): os.makedirs(uncompress_path)
        patoolib.extract_archive(compress_path, outdir=uncompress_path)
        result = 0
    except Exception as e:
        try:
            shutil.rmtree(uncompress_path)
        except:
            pass
        logging.error(str(e))
    finally:
        return result

def pdf2img(pdf_path, to_path):
    '''
    # 将PDF转化为图片
    pdf_path pdf文件的路径
    to_path 保存文件夹目录
    zoom_x x方向的缩放系数
    zoom_y y方向的缩放系数
    rotation_angle 旋转角度
    '''
    try:
        import fitz
        doc = fitz.open(pdf_path)
        print("PDF转图片，任务文件：%s" % pdf_path)

        ts = now()
        for pg in range(doc.pageCount):
            print("\r共%s页,正在转换第%s/%s张" % (doc.pageCount, pg+1, doc.pageCount), end="")
            page = doc[pg]
            rotate = int(0)     # 旋转角度
            # 分辨率缩放系数；x=4，y=4，得到的图片分辨率尺寸是原来的16倍
            zoom_x = 2.0
            zoom_y = 2.0
            trans = fitz.Matrix(zoom_x, zoom_y).prerotate(rotate)
            pm = page.get_pixmap(matrix=trans, alpha=False)
            save_name = '{:01}.png'.format(pg+1)
            pm.save(os.path.join(to_path, save_name))
        print()
        print("耗时:%s秒" % str(now() - ts))
    except Exception as e:
        print(str(e))

def read_pd(filepath_or_io,
            type: str="execl",
            **kwargs):
    """
    读取execl文件
    :param filepath:
    :return:

    io，Excel的存储路径
    sheet_name，要读取的工作表名称
    header， 用哪一行作列名
    names， 自定义最终的列名
    index_col， 用作索引的列
    usecols，需要读取哪些列
    squeeze，当数据仅包含一列
    engine：编写要使用的引擎“openpyxl”或“xlsxwriter”。 您还可以通过选项io.excel.xlsx.writer，io.excel.xls.writer和io.excel.xlsm.writer进行设置。
    converters ，强制规定列数据类型
    skiprows，跳过特定行
    nrows ，需要读取的行数
    skipfooter ， 跳过末尾n行

    """
    import pandas
    mp = {
        "execl": pandas.read_excel,
        # "xls": pandas.read_excel, # openpyxl does not support the old .xls file format
        "xlsx": pandas.read_excel,
        "xlsm": pandas.read_excel,
        "csv": pandas.read_csv,
        "json": pandas.read_json,
        "table": pandas.read_table
    }

    if type in mp:
        df = mp[type](filepath_or_io, **kwargs)
    else:
        raise Exception("require pecify format: %s" % to_json(list(mp.keys())))

    return df

def image2thumb(image_path,
                save_path,
                save_name,
                width: int=200,
                height: int=100):
    """
    图片转缩略图
    :param image_path:
    :param out_path:
    :return:
    """
    from PIL import Image
    __, filename = os.path.split(image_path)
    filename, default_format = os.path.splitext(filename)
    save_name, format = os.path.splitext(save_name)
    format = format if "." not in format else format.lstrip(".")
    format = format or default_format.lstrip(".")

    save_path = os.path.join(save_path.rstrip("/"), save_name + "." + format.lower())

    im = Image.open(image_path)
    im.thumbnail((width, height))
    im.save(save_path)

def txt2execl(from_path_or_lines, out_path, sheet_name="Sheet1"):
    """
    1、txt固定格式文本转存execl
    2、固定长度的二维数组转存execl:

    txt样例:
    列1  列2  列3
    a    b    c
    x    y    z
    o    q    k

    二维数组样例：
     [
        [列1  列2  列3],
        [a    b    c],
        [x,   y,   z],
        [o    q    k]
    ]

    Excel结果：
    列1  列2  列3
    a    b    c
    x    y    z
    o    q    k
    :param from_path_or_lines:
    :param out_path:
    :param sheet_name:
    :return:
    """

    from openpyxl import Workbook, load_workbook

    if isinstance(from_path_or_lines, list):
        contents = from_path_or_lines
    else:
        contents = []
        with open(from_path_or_lines, encoding="utf-8", errors='ignore') as fp:
            for line in fp.readlines():
                contents.append(line.split())

    if not os.path.exists(out_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        workbook.save(out_path)

    execl = load_workbook(out_path)
    # execl = execl.create_sheet(sheet_name)
    sheet = execl[sheet_name]
    for i in contents:
        sheet.append(i)
    execl.save(out_path)

def gen_imgcode():
    """
    生成图片验证码
    """
    '''
            图片验证码（完成）
            GET route
            -------uuid
            :return:
            '''
    from six import BytesIO
    from PIL import Image, ImageDraw, ImageFont
    # 定义变量，用于画面的背景色、宽、高
    bgcolor = (255, 255, 255)
    width = 100
    height = 25
    # 创建画面对象
    im = Image.new('RGB', (width, height), bgcolor)
    # 创建画笔对象
    draw = ImageDraw.Draw(im)
    # 调用画笔的point()函数绘制噪点
    for i in range(0, 100):
        xy = (random.randrange(0, width), random.randrange(0, height))
        fill = (12, 131, 250)
        draw.point(xy, fill=fill)
    # 定义验证码的备选值
    str1 = 'ABCD123EFGHIJK456LMNOPQRS789TUVWXYZ0'
    # 随机选取4个值作为验证码
    rand_str = ''
    for i in range(0, 4):
        rand_str += str1[random.randrange(0, len(str1))]
    # 构造字体
    font = ImageFont.truetype('FreeMono.ttf', 23)
    # 构造字体颜色
    fontcolor = (12, 131, 250)
    # 绘制4个字
    draw.text((5, 2), rand_str[0], font=font, fill=fontcolor)
    draw.text((25, 2), rand_str[1], font=font, fill=fontcolor)
    draw.text((50, 2), rand_str[2], font=font, fill=fontcolor)
    draw.text((75, 2), rand_str[3], font=font, fill=fontcolor)
    # 释放画笔
    del draw
    # 存入session，用于做进一步验证
    # request.session['verifycode'] = rand_str
    # 内存文件操作
    buf = BytesIO()
    # 将图片保存在内存中，文件类型为png
    im.save(buf, 'png')
    # 将内存中的图片数据返回给客户端，MIME类型为图片png
    return buf.getvalue(), 'image/png'