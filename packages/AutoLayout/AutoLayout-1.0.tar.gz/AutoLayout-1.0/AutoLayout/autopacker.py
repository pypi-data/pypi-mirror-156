import gdspy
from rectpack import newPacker
import warnings
from .numCodeGenerator import morseNumCodeGenerator
from openpyxl import Workbook
import string
from copy import deepcopy
from random import shuffle

def region_autopacker(region_list, mandatory_list, optional_list = []):
  """Takes list of cells from mandatory and optional lists and packs them onto the cells in region list

  Args:
    region_list: the list of cells that contains the regions (must be rectangle only) 
                 that needs to be packed onto, one region per cell only. This has to be a list
    mandatory_list: the mandatory list of cells that needs to be packed onto the regions
    optional_list: defaul is an empty list. The prioritized optional list of cells that will 
                   be packed if there is extra room in the regions
    region_layer: layer of the region, default is 0
    region_datatype: datatype of the region, default is 0


  Returns:
    List of cells(regions) that already packed

  Raises: 
    AssertionError: Can not pack the mandatory list onto the regions
    ValueError: The region.name() can only have ONE rectangle!
  """

  #Preparation

  packer = newPacker(rotation=False)
  devices = mandatory_list + optional_list
  bid, rid = 0,0
  bid_object = {} # a dictionary that key is the bid and value is the object -- [bid]:[object]
  rid_object = {} # a dictionary that key is the rid and value is the object -- [rid]:[object]
  
  available_area, packing_area, potential_devices = 0,0,0
  devices = mandatory_list + optional_list

  #Calculate avaiable area, packing area and potential devices
  for region in region_list:
    region_dim = region.get_buffer_bounding_box()
    w = region_dim[1][0] - region_dim[0][1]
    h = region_dim[1][1] - region_dim[0][1]
    available_area += w * h

  while packing_area <= available_area and potential_devices < len(devices):
    if devices[0].dev_type == True:
      device_dim = devices[potential_devices].get_buffer_bounding_box()

    else:
      device_dim = devices[potential_devices].get_region_bounding_box()
    w = device_dim[1][0] - device_dim[0][1]
    h = device_dim[1][1] - device_dim[0][1]
    packing_area += w * h
    potential_devices += 1
  #Roughly check the potential devices with mandatory list
  if potential_devices < len(mandatory_list):
    raise AssertionError(f'Can not pack the {mandatory_list} onto the {region_list}.\n')

  #Start packing
  while True: #Do-while loop in Python
    #Add bin
    for region in region_list:
      region_dim = region.get_buffer_bounding_box()
      w = region_dim[1][0] - region_dim[0][1]
      h = region_dim[1][1] - region_dim[0][1]
      packer.add_bin(w,h, bid = bid)
      bid_object[bid] = region
      bid += 1

    #Add devices
    for device in devices[:potential_devices]:
      if device.dev_type == True:
        device_dim = device.get_buffer_bounding_box()
      else:
        device_dim = device.get_region_bounding_box()
      w = device_dim[1][0] - device_dim[0][1]
      h = device_dim[1][1] - device_dim[0][1]
      packer.add_rect(w,h,rid = rid)
      rid_object[rid] = device
      rid += 1

    #Pack
    packer.pack()
    #Check and stop the loop
    if len(packer.rect_list()) == potential_devices:
      break
    #If false, reset and ready for the new loop
    packer = newPacker(rotation=False)
    bid_object={}
    rid_object={}
    bid, rid = 0,0
    potential_devices -= 1
  #Check for mandatory list and raise error
  if len(packer.rect_list()) < len(mandatory_list): 
    raise AssertionError('Can not pack the mandatory list onto the regions.\n')

  #Warn the user if one of the region is not used
  if len(packer) < len(region_list):
    warnings.warn('One or more regions were not used. Those regions will not be packed further')

  # Records the origin and add references of all the devices in the regions
  regions = []
  for abin in packer:
    a_region = bid_object[abin.bid]
    buffer = a_region.buffer_size
    for rect in abin:
      a_region.add(rid_object[rect.rid])
      rid_object[rect.rid].set_origin(rect.x + buffer,rect.y + buffer)      
    regions.append(a_region)
  return regions

up_name = '1'
def finalIntegration(object):
  """Renames the cell in order and adds CellReference in gdspy to the region cell

  Args:
    object: the master cell, usually its the dice object

  """
  global up_name
  global count
  if object.dev_type == False:
      count = 1
      for item in object.references:
          item.set_numCode(count)
          item.cell.name = up_name + ',' + str(count)
          item.id = '(' + item.cell.name + ')'
          object.cell.add(gdspy.CellReference(item.cell, origin=(item.x,item.y)))
          count = count + 1

      count = 1
      for item in object.references:
          up_name = item.cell.name
          finalIntegration(item)
          count = count + 1

class autopacker:
  """
  A packer used to pack the objects (devices, regions) in a hierachy way
  """
  def __init__(self):
    pass

  def set_device(self, devices_list, randomOptional = False):
    """
    Sets the devices list into the packer, each element of the list is a tuple that contains
    (device, mandatory multiple times, optional multiple times)

    Args:
      devices_list (list): [(device, mandatory times, optional times)], optional list is prioritized by the order of the devices list.
      randomOptional (bool): set this one True will shuffle the optional devices.
    """
    self.dev_type = []
    self.mandatory = []
    self.optional = []

    self.mandatory_orig=[]
    self.optional_orig=[]
    self.dev_type_orig=[]

    for each in devices_list:
      device = each[0]
      self.dev_type_orig.append(device.typeLabel)
      for i in range(each[1]):
        self.mandatory_orig.append(device.copy())
      for i in range(each[2]):
        self.optional_orig.append(device.copy())   
    
    #Shuffle optional list
    if randomOptional:
      shuffle(self.optional_orig)
    self.device_reset()

  def device_reset(self):
    self.dev_type=deepcopy(self.dev_type_orig)
    self.optional=deepcopy(self.optional_orig)
    self.mandatory=deepcopy(self.mandatory_orig)

  def set_region(self, region_list):
    """Sets the regions list into the packer, 
    each element will be packed by the element behind it in this list
    
    Args:
      region_list: [[dice],[region1,region2],...]

    Raises:
      TypeError if each element of the list is not a list itself
    """

    for regionsList in region_list:
      if type(regionsList) != list:
        raise TypeError(f'This {regionsList} is not a list')
    if len(region_list[0]) != 1:
      raise TypeError('There must be ONE AND ONLY ONE dice to pack everything onto')
    self.originalRegions=deepcopy(region_list)

  def region_reset(self):
    region_listCopy=[]
    for alist in self.originalRegions:
      newRegions=[]
      for area in alist:
        newRegions.append(area.copy())
      region_listCopy.append(newRegions)
    self.regions = region_listCopy
    self.regions.reverse()

  def potential_pack(self):
    """Checks the potential devices can be packed onto region lists
    
    Raises:
      AssertionError: Can not pack the mandatory list onto the regions

    Return:
      The potential devices can be pack onto the region lists
    """

    available_area, packing_area, potential_devices = 0,0,0

    devices = self.mandatory + self.optional

    #Calculate avaiable area, packing area and potential devices
    for region in self.regions[0]:
      region_dim = region.cell.get_bounding_box()
      w = region_dim[1][0] - region_dim[0][1]
      h = region_dim[1][1] - region_dim[0][1]
      available_area += w * h

    while packing_area <= available_area and potential_devices < len(devices):
      device_dim = devices[potential_devices].cell.get_bounding_box()
      deviceBuffSize=devices[potential_devices].devBufferSize
      devNumCodeSize=devices[potential_devices].numcode_size
      devNumCodeOffsets=devices[potential_devices].numcode_offset
      devEtchSpacing=devices[potential_devices].numcode_etchSpacing

      w = (device_dim[1][0] - device_dim[0][1])+2*deviceBuffSize + 10*devEtchSpacing + 2*devNumCodeSize
      h = device_dim[1][1] - device_dim[0][1]+2*deviceBuffSize + devNumCodeSize + devNumCodeOffsets

      #print(packing_area/available_area)
      packing_area += w * h
      potential_devices += 1
    #print(available_area, packing_area)
    #Roughly check the potential devices with mandatory list
    if potential_devices < len(self.mandatory):
      raise AssertionError('Can not pack the mandatory list onto the regions.\n')
    
    return potential_devices

  def autoLayoutPostProcessing(self, max_numCode, potential_devices):
    """
    Runs autoLayoutPostProcessor of each device or region

    Args:
      max_numCode: the max num code to reserve room for actual num code after everything being pack
      potential_devices: the potential devices that gonna be packed. This number will be used to limit the process, if not 
                        the process will run for the entire device list, which can take a lot of time
                       
    """
    self.optional = self.optional[:(potential_devices-len(self.mandatory))]
    self.devices = self.mandatory + self.optional
    for dev in self.devices:
      dev.autolayoutPostProcessor(max_numCode)
    for regionsList in self.regions:
      region_maxNumCode = len(regionsList)
      region_maxNumCode = int(region_maxNumCode / 10) * 10 + 9
      for region in regionsList:
        if region.numcode_infor:
          region.autolayoutPostProcessor(region_maxNumCode)
        else:
          region.autolayoutPostProcessor()

  def pack(self):
    """Packs everything in a hierarchal way"""

    #Reset the device and region lists
    self.region_reset()
    self.device_reset()
    
    #Set a dummie numcode to make room for actuall numcode that 
    # will be assigned after packing

    potential_devices = self.potential_pack()
    max_numCode = int(potential_devices / len(self.regions[0]) )
    max_numCode = int(max_numCode / 10) * 10 + 9
    self.autoLayoutPostProcessing(max_numCode, potential_devices)
    if self.mandatory:
      regions = region_autopacker(self.regions[0],self.mandatory,self.optional)
    else:
      raise ValueError('There is no device to pack')
    for i in range(len(self.regions)-1):
      regions = region_autopacker(self.regions[i+1], regions)
      self.dice = regions[0]

  def to_layout(self, filename):
    """Makes .gds layout for the autopacker
    
    Args:
      filename: the name of the output file (no extention .gds at the end)

    Return:
      a file with the 'filename'.gds in the same directory
    """
    
    self.dice.cell.name = '1'
    finalIntegration(self.dice)
    lib = gdspy.GdsLibrary()
    lib.add(self.dice.cell, include_dependencies=True)
    lib.write_gds(filename + '.gds')
    return lib
  
  def to_excel(self, filename):
    """Makes .xlsx excel file with each worksheet is a type of device in their information

    Args:
      filename: the name of the output file (no extention .xlsx at the end)

    Return:
      a excel file in the same directory
    """
    packed_devices = [device for device in self.devices if device.id]
    sorted_devices = []

    #Eliminate duplication in dev_type
    tempDevType=set(self.dev_type)
    tempDevType=list(tempDevType)

    #Sort devices by type label
    for devName in tempDevType:
      l = [device for device in packed_devices if device.typeLabel == devName]
      sorted_devices.append(l)
    
    #Create excel file
    wb = Workbook()
    #Delete the blank sheet
    ws = wb.active
    wb.remove(ws)

    #Create sheet and fill
    for typeDevice in sorted_devices:
      # print(typeDevice)
      representative_dev = typeDevice[0]

      #Make a sheet for each type of device
      ws = wb.create_sheet(representative_dev.typeLabel)

      #Create the first line
      ws['A1'] = 'Device Label'
      ws['B1'] = 'Dev ID'
      ws['C1'] = 'Local location'
      columns = 4
      for parameter in representative_dev.parameters:
        d = ws.cell(row=1, column=columns)
        d.value = str(parameter)
        columns += 1
      
      #Create the follow up lines with device information
      row = 2
      for device in typeDevice:
        ws[f'A{row}'] = device.typeLabel
        ws[f'B{row}'] = device.id
        ws[f'C{row}'] = f'({int(device.x)}, {int(device.y)})'
        columns= 4
        for parameter in device.parameters.values():
          d = ws.cell(row=row, column=columns)
          d.value = str(parameter)
          columns += 1
        row += 1
      
    wb.save(f'{filename}.xlsx')
    

    

        
      



